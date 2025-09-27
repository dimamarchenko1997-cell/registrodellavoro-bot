import os
import asyncio
import calendar
from datetime import datetime
from math import radians, sin, cos, sqrt, atan2
from aiogram import Bot, Dispatcher, F, types
from aiogram.types import Message, CallbackQuery, ReplyKeyboardMarkup, KeyboardButton
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook, load_workbook
import aioschedule
from dotenv import load_dotenv
import logging
from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse

# ---------------- CONFIG ----------------
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN") or "PASTE_YOUR_TOKEN_HERE"
EXCEL_FILE = "registro_lavoro.xlsx"
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()
logging.basicConfig(level=logging.INFO)

# ---------------- Constants ----------------
WORK_LOCATIONS = {"Ufficio Centrale": (45.602129, 9.248768)}
MAX_DISTANCE_METERS = 200

# ---------------- FSM ----------------
class RegistroForm(StatesGroup):
    waiting_ingresso_location = State()
    waiting_uscita_location = State()

class PermessiForm(StatesGroup):
    waiting_for_start = State()
    waiting_for_end = State()
    waiting_for_reason = State()

# ---------------- Excel ----------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Registro"
        ws1.append(["Data", "Utente", "Ingresso ora", "Posizione ingresso", "Uscita ora", "Posizione uscita"])
        ws2 = wb.create_sheet("Permessi")
        ws2.append(["Data richiesta", "Utente", "Dal", "Al", "Motivo"])
        wb.save(EXCEL_FILE)

def save_ingresso(user: types.User, time, location_name):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Registro"]
        today = datetime.now().strftime("%d.%m.%Y")
        user_id = f"{user.full_name} | {user.id}"
        # Controlla se esiste già un ingresso per oggi
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == today and ws.cell(row, 2).value == user_id:
                logging.warning(f"Ingresso già registrato per {user_id} oggi.")
                return False
        ws.append([today, user_id, time, location_name, "", ""])
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        logging.error(f"Errore durante il salvataggio dell'ingresso: {e}")
        return False

def save_uscita(user: types.User, time, location_name):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Registro"]
        today = datetime.now().strftime("%d.%m.%Y")
        user_id = f"{user.full_name} | {user.id}"
        for row in range(ws.max_row, 0, -1):
            if ws.cell(row, 1).value == today and ws.cell(row, 2).value == user_id and not ws.cell(row, 5).value:
                ws.cell(row, 5, time)
                ws.cell(row, 6, location_name)
                wb.save(EXCEL_FILE)
                return True
        logging.warning(f"Nessun ingresso trovato per {user_id} oggi.")
        return False
    except Exception as e:
        logging.error(f"Errore durante il salvataggio dell'uscita: {e}")
        return False

def save_permesso(user: types.User, start_date, end_date, reason):
    try:
        # Validazione date
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        if end_dt < start_dt:
            raise ValueError("La data di fine deve essere successiva o uguale alla data di inizio.")
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Permessi"]
        today = datetime.now().strftime("%d.%m.%Y %H:%M")
        user_id = f"{user.full_name} | {user.id}"
        ws.append([today, user_id, start_date, end_date, reason])
        wb.save(EXCEL_FILE)
        return True
    except ValueError as ve:
        logging.error(f"Errore di validazione: {ve}")
        return False
    except Exception as e:
        logging.error(f"Errore durante il salvataggio del permesso: {e}")
        return False

# ---------------- Location check ----------------
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    dlat, dlon = radians(lat2 - lat1), radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return R * 2 * atan2(sqrt(a), sqrt(1 - a))

def check_location(lat, lon):
    for name, (wlat, wlon) in WORK_LOCATIONS.items():
        if haversine(lat, lon, wlat, wlon) <= MAX_DISTANCE_METERS:
            return name
    return None

# ---------------- Keyboards ----------------
main_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Ingresso")],
        [KeyboardButton(text="Uscita")],
        [KeyboardButton(text="Richiesta permessi")]
    ],
    resize_keyboard=True
)

# ---------------- Calendar ----------------
def mese_nome(month: int) -> str:
    mesi = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
    return mesi[month - 1]

def build_calendar(year: int, month: int, phase: str):
    kb = InlineKeyboardBuilder()
    today = datetime.now()
    giorni = ["Lu", "Ma", "Me", "Gi", "Ve", "Sa", "Do"]
    
    # Prima riga: 3 celle con frecce più piccole e mese/anno più grande
    # Riduci spazi per frecce (più piccole), aggiungi più spazi per mese (più grande)
    kb.button(text="⬅️", callback_data=f"perm:{phase}:nav:{year}:{month}:prev")
    kb.button(text=f"     {mese_nome(month)} {year}     ", callback_data="ignore")
    kb.button(text="➡️", callback_data=f"perm:{phase}:nav:{year}:{month}:next")
    
    # Seconda riga: giorni della settimana
    for g in giorni:
        kb.button(text=g, callback_data="ignore")
    
    # Numeri del mese
    for week in calendar.monthcalendar(year, month):
        for day in week:
            if day == 0:
                kb.button(text=" ", callback_data="ignore")
            else:
                text_day = f"🔵{day}" if day == today.day and month == today.month and year == today.year else str(day)
                kb.button(text=text_day, callback_data=f"perm:{phase}:day:{year}:{month}:{day}")
    
    # Imposta le larghezze delle righe: prima riga 3, poi 7 per giorni, e 7 per ciascuna settimana
    kb.adjust(3, 7, 7, 7, 7, 7, 7)
    
    return kb.as_markup()

# ---------------- Handlers ----------------
@dp.message(F.text == "/start")
async def start_handler(message: Message):
    await message.answer("Benvenuto! Scegli un'opzione:", reply_markup=main_kb)

@dp.message(F.text == "Ingresso")
async def ingresso_start(message: Message, state: FSMContext):
    await state.set_state(RegistroForm.waiting_ingresso_location)
    kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📍 Invia posizione", request_location=True)]],
        resize_keyboard=True
    )
    await message.answer("Invia la tua posizione per registrare l'ingresso:", reply_markup=kb)

@dp.message(RegistroForm.waiting_ingresso_location, F.location)
async def ingresso_location(message: Message, state: FSMContext):
    loc = message.location
    location_name = check_location(loc.latitude, loc.longitude)
    if location_name:
        now = datetime.now().strftime("%H:%M")
        if save_ingresso(message.from_user, now, location_name):
            await message.answer("✅ Ingresso registrato!", reply_markup=main_kb)
        else:
            await message.answer("❌ Ingresso già registrato per oggi.", reply_markup=main_kb)
    else:
        await message.answer("❌ Non sei in un luogo autorizzato.", reply_markup=main_kb)
    await state.clear()

@dp.message(F.text == "Uscita")
async def uscita_start(message: Message, state: FSMContext):
    await state.set_state(RegistroForm.waiting_uscita_location)
    kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📍 Invia posizione", request_location=True)]],
        resize_keyboard=True
    )
    await message.answer("Invia la tua posizione per registrare l'uscita:", reply_markup=kb)

@dp.message(RegistroForm.waiting_uscita_location, F.location)
async def uscita_location(message: Message, state: FSMContext):
    loc = message.location
    location_name = check_location(loc.latitude, loc.longitude)
    if location_name:
        now = datetime.now().strftime("%H:%M")
        if save_uscita(message.from_user, now, location_name):
            await message.answer("✅ Uscita registrata!", reply_markup=main_kb)
        else:
            await message.answer("❌ Nessun ingresso trovato per oggi.", reply_markup=main_kb)
    else:
        await message.answer("❌ Non sei in un luogo autorizzato.", reply_markup=main_kb)
    await state.clear()

# ---- PERMESSI ----
@dp.message(F.text == "Richiesta permessi")
async def permessi_start(message: Message, state: FSMContext):
    await state.set_state(PermessiForm.waiting_for_start)
    now = datetime.now()
    await message.answer("📅 Seleziona data di inizio:", reply_markup=build_calendar(now.year, now.month, "start"))

@dp.callback_query(F.data.startswith("perm:"))
async def perm_calendar_handler(cb: CallbackQuery, state: FSMContext):
    parts = cb.data.split(":")
    if len(parts) < 3:
        await cb.answer()
        return
    phase, kind = parts[1], parts[2]
    if kind == "nav":
        year = int(parts[3])
        month = int(parts[4])
        direction = parts[5]
        if direction == "prev":
            if month == 1:
                month = 12
                year -= 1
            else:
                month -= 1
        else:
            if month == 12:
                month = 1
                year += 1
            else:
                month += 1
        await cb.message.edit_reply_markup(reply_markup=build_calendar(year, month, phase))
        await cb.answer()
        return
    if kind == "day":
        year = int(parts[3])
        month = int(parts[4])
        day = int(parts[5])
        selected = f"{year}-{month:02d}-{day:02d}"
        if phase == "start":
            await state.update_data(start_date=selected)
            await state.set_state(PermessiForm.waiting_for_end)
            await cb.message.edit_text(
                f"📅 Inizio selezionato: {selected}\nSeleziona la data di fine:",
                reply_markup=build_calendar(year, month, "end")
            )
        elif phase == "end":
            await state.update_data(end_date=selected)
            await state.set_state(PermessiForm.waiting_for_reason)
            await cb.message.edit_text(
                f"📅 Fine selezionata: {selected}\nOra scrivi il motivo del permesso:"
            )
        await cb.answer()

@dp.message(PermessiForm.waiting_for_reason)
async def permessi_reason(message: Message, state: FSMContext):
    data = await state.get_data()
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    reason = message.text or ""
    if save_permesso(message.from_user, start_date, end_date, reason):
        await message.answer("✅ Permesso registrato!", reply_markup=main_kb)
    else:
        await message.answer("❌ Errore nella registrazione del permesso (data non valida o altro problema).", reply_markup=main_kb)
    await state.clear()

# ---------------- Scheduler ----------------
async def notify_missing_ingresso():
    today = datetime.now().strftime("%d.%m.%Y")
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Registro"]
        # Qui dovresti avere una lista di utenti attesi. Per ora, assumiamo di loggare solo.
        # Per inviare notifiche, avresti bisogno di una lista di user_id.
        registered_users = {ws.cell(row, 2).value for row in range(2, ws.max_row + 1) if ws.cell(row, 1).value == today}
        logging.info(f"[NOTIFY] Utenti con ingresso oggi ({today}): {registered_users}")
        # Esempio: per inviare notifiche, loop su user_id noti e controlla se non in registered_users
        # await bot.send_message(user_id, "Ricorda di registrare l'ingresso!")
    except Exception as e:
        logging.error(f"Errore nella notifica: {e}")

async def scheduler():
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(2)

async def on_startup():
    init_excel()
    days = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    for day in days:
        getattr(aioschedule.every(), day).at("09:00").do(notify_missing_ingresso)
    asyncio.create_task(scheduler())
    logging.info("🚀 Bot avviato con webhook su Render")

# ---------------- FastAPI Setup ----------------
app = FastAPI()

@app.on_event("startup")
async def startup_event():
    await on_startup()

@app.post("/webhook")
async def webhook(request: Request):
    try:
        # Aggiungi context={"bot": bot} per la validazione corretta (richiesto da aiogram/pydantic)
        update = types.Update.model_validate(await request.json(), context={"bot": bot})
        await dp.feed_update(bot=bot, update=update)
        return {"ok": True}  # Telegram si aspetta una risposta semplice con status 200
    except Exception as e:
        logging.error(f"Errore nel webhook: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)  # Per debug, ma non obbligatorio
@app.get("/")
async def health_check():
    return "Bot is running"

# ---------------- Main ----------------
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
